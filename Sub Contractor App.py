import sys
import os 
import openpyxl 
import datetime
import shutil

#Globals
paymentNoCell = "H13"
previousPaymentForwardCell = "J20"
today = datetime.date.today()
date = today.strftime("%d/%m/%Y")


#System Setup
clear = lambda: os.system('cls')
applicationPath = os.environ['APPLICATION_PATH']


class PaymentCertificate():
    """"Generate a new payment certificate"""

    def __init__(self) -> None:

        PaymentCertificate.selectedContract = Contract
        PaymentCertificate.selectedContract.SelPayment(self)

        self.SelectCompany()

    def SelectCompany(self):
        
        filePath = self.selectedContract.filePath+"/Sub-Contractors/Payments"
        directoryNames = []

        for (dirpath, dirnames, filenames) in os.walk(filePath):
                directoryNames.extend(dirnames)

        choices = ""
        for index in range(len(directoryNames)):
            choices += (str(index)+" - "+directoryNames[index]+"\n")

        choices += (str(len(directoryNames))+" - New Company \n")
        print(choices)

        companySelection = int(input("Selection: "))
        clear()

        try:
            
            filePath = filePath+"/"+directoryNames[companySelection]

        except:

            filePath
                
        PaymentCertificate.selectedContract.filePath = filePath

        if companySelection == len(directoryNames):
            
            PaymentCertificate.MakeNewPayment()

        else:
            PaymentCertificate.selectedName = directoryNames[companySelection]
            self.MakeExistingPayment()    

    def MakeNewPayment():

        
        print("New Company Name")
        projectNo = ((PaymentCertificate.selectedContract.filePath).rsplit('/' )[2])[0:5]
        projectName = ((PaymentCertificate.selectedContract.filePath).rsplit('/')[2])[5:]
        
        companyName = input("Company Name: ")
        PaymentCertificate.selectedName = companyName

        PaymentCertificate.excelPath = PaymentCertificate.selectedContract.filePath + "/" + companyName + "/" + companyName + ".xlsx"
        os.makedirs(os.path.dirname(PaymentCertificate.excelPath), exist_ok=True)
        shutil.copyfile(applicationPath + r"\Data\Subcontractor Certificate Template.xlsx", PaymentCertificate.excelPath)

        wb = openpyxl.load_workbook(PaymentCertificate.excelPath)
        previousSheet = wb["001"]
        wb.copy_worksheet(previousSheet)
        wb.remove(previousSheet)
        newSheet = wb["001 Copy"]
        newSheet.title = "000"

        newSheet["A12"] = companyName
        newSheet["H5"] = projectNo
        newSheet["H7"] = projectName
        newSheet["H13"] = 0

        newSheet["H17"] = input("Input QS Name: ")
        clear()

        print("Input "+companyName+" Address:")
        for index in range(5):
            newSheet["A"+str(13+index)] = input("Line "+str(index+1)+" ")
        clear()

        print("Select CIS Status: \n1 - Gross\n2 - 20%\n3 - 30%")
        selection = int(input("Input: "))
        if selection == 1:
            newSheet["H15"] = "GROSS"

        elif selection == 2:
            newSheet["H15"] = "20%"

        elif selection == 3:
            newSheet["H15"] = "30%"

        saved = False
        
        while saved == False:

            try:
                wb.save(PaymentCertificate.excelPath)
                saved = True
                
            except:
                input("File open, please close then press enter to continue")
        
        PaymentCertificate.selectedContract.filePath = PaymentCertificate.selectedContract.filePath + "/" + companyName
        PaymentCertificate.MakeExistingPayment()

    def MakeExistingPayment():

        companyName = PaymentCertificate.selectedName 
        print("Making New Payment To "+companyName)
        
        #Open Excel Doc
        excelPath = PaymentCertificate.excelPath
        wb = openpyxl.load_workbook(excelPath)
        wbData = openpyxl.load_workbook(filename = excelPath,data_only=True)

        #Find Previous Payment
        highestValue = 0

        for sheet in wb.worksheets:
            value = sheet[paymentNoCell].value
            if value > highestValue:
                highestValue = value

        #DONT FORGET TO MATCH SHEET# TO PAYMENT#
        paymentNo = highestValue+1
        
        #Define Previous Payment Sheet
        previousSheetName = '{0:03}'.format(paymentNo-1)
        previousSheet = wb[previousSheetName]
        wb.copy_worksheet(previousSheet)

        #Define New Payment Sheet 
        newSheetName = '{0:03}'.format(paymentNo)
        newSheet = wb[previousSheetName+" Copy"]
        newSheet.title = newSheetName
        newSheet = wb[newSheetName]

        #Define New Payment Sheet Data
        previousSheetData = wbData[previousSheetName]

        # Update Amount(£), Payment No., Previously Certified, Payment Bought Forward and Date
        
        newSheet["E48"] = "=SUM(J22:J47)"
        
        newSheet[paymentNoCell] = paymentNo

        newSheet["J62"] = previousSheetData["J60"].value
        newSheet["J20"] = previousSheetData["J49"].value

        newSheet["C58"] = date

        #Formatting

        for i in range(22, 47):
            newSheet["F"+str(i)] = ""
            newSheet["A"+str(i)] = f"""=IF(F{str(i)}<>"",COUNT($F$22:F{str(i)})+1,"")"""

        img = openpyxl.drawing.image.Image(applicationPath+r'\Images\CompanyLogo.png')
        img.anchor = 'A1'
        img.width =  360
        img.height = 60
        newSheet.add_image(img)
        
        #Inputs

        #ID and Date

        print("Input Invoice ID")
        print("Previous ID: "+str(previousSheetData["C48"].value))
        newSheet["C48"] = input("Invoice ID: ")
        clear()

        print("Input Invoice Date (00/00/00)")
        newSheet["H48"] = input("Date: ")
        clear()

        #Due Date

        nextFriday = today + datetime.timedelta( ((4-today.weekday()) % 7) +7 )
        print("Select or Input Due Date")
        print("1 - Friday ("+str((nextFriday- datetime.timedelta(days=7)).strftime("%d/%m/%Y"))+")")
        print("2 - Next Friday ("+str(nextFriday.strftime("%d/%m/%Y"))+")")
        print("Input - Custom Date (00/00/00)")
        selection = input("Selection or Custom Date: ")

        if selection == "1":
            newSheet["C64"] = str((nextFriday- datetime.timedelta(days=7)).strftime("%d/%m/%Y"))

        elif selection == "2":
            newSheet["C64"] = str(nextFriday.strftime("%d/%m/%Y"))
        else:
            print("Custom date "+selection+" recieved")
            newSheet["C64"] = selection

        #Sign?
        print("Sign? \n1 - Yes\n2 - No")
        selection = int(input("Selection: "))

        if selection == 1:
            img = openpyxl.drawing.image.Image(applicationPath+r'\Images\Signature\JohnSmith.PNG')
            img.anchor = 'C55'
            
            img.width =  135
            img.height = 50
            newSheet.add_image(img)

        #Save
        saved = False
        while saved == False:

            try:
                wb.save(excelPath)
                saved = True

            except:
                input("File open, please close then press enter to continue")

        print("Saved = "+str(saved))
        os.startfile(excelPath)
     
class Contract():   
    "To be moved to a utils folder for other apps"
    

    def SelPayment(self):
        #Variables

        def Login():

            index = -1
            print("Select Name")

            for (dirpath, dirnames, filenames) in os.walk(applicationPath+"/Data/Memory"):
                for fileName in filenames:
                    index += 1
                    print(str(index)+" - "+fileName)

                print(str(len(filenames))+" - New")
                selection = int(input("Selection: "))

                if selection == len(filenames):
                    loginPath = applicationPath+"/Data/Memory/"+input("Input name ")+".txt"
                    f = open(loginPath, "x")
                    return loginPath

                else:
                    return applicationPath+"/Data/Memory/"+filenames[selection]


        loginPath = Login()
        filePath = "\\Contracts Folder"
        memory = open(loginPath, "r")
        
        #Recent Payment Choices
        selectionMemory = memory.read()
        selectionMemory = selectionMemory.split("\n")

        if len(selectionMemory) > 3:
            selectionMemory.pop()

        choices = ""

        for index in range(len(selectionMemory)):
                choices += (str(index)+" - "+selectionMemory[index].rsplit('/', 1)[-1]+"\n")

        choices += str(len(selectionMemory))+" - Other"
        print("Recently Selected Contracts")
        print(choices)
        selection = int(input("Selection: "))
        print("Other selected")
        clear()

        #All contracts choices
        if selection == len(selectionMemory):
            
            for i in range(2):

                if i == 2:
                    selectionMemory.insert(0,filePath)
                    with open(applicationPath+"\Data\contractMemory.txt", "w") as file:
                        file_lines = "\n".join(selectionMemory)
                        file.write(file_lines)
                    
                directoryNames = []

                for (dirpath, dirnames, filenames) in os.walk(filePath):
                    directoryNames.extend(dirnames)
                    break
                
                if i == 0:
                    filteredDirectoryNames = [k for k in directoryNames if 'Contracts' in k]
                    directoryNames = filteredDirectoryNames

                choices = ""
                for index in range(len(directoryNames)):
                    choices += (str(index)+" - "+directoryNames[index]+"\n")

                print(choices)
                companySelection = int(input("Selection: "))

                try:
                    filePath = filePath+"/"+directoryNames[companySelection]

                except:
                    x=1
        else:
            filePath = selectionMemory[selection]
            
        print(filePath)
        Contract.filePath = filePath
             

def Main():

    PaymentCertificate()


if __name__ == '__main__':
    sys.exit(Main())




